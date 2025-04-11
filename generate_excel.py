from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy
import io

def generate_excel_file(productos, template_path, info=None):
    """
    Encabezado (igual que antes) => D13, D14, etc.
    Tabla de productos:
      - B => Ítem
      - C..D => Descripción (merge)
      - E => Unidad
      - F => Cant
      - G => Precio Unit
      - H => Precio Parcial
      - K..L => Marca/Modelo (merge)
    Luego sumamos la col H y creamos el footer con insert_footer_section.
    """
    if info is None:
        info = {}

    wb = load_workbook(template_path)
    ws = wb.active

    # 1) Rellenar encabezado
    ws["D13"].value = info.get('cliente', "")
    ws["D14"].value = info.get('solicitante', "")
    ws["D15"].value = info.get('email', "")
    ws["D16"].value = info.get('referencia', "")

    ws["J13"].value = info.get('ruc', "")
    ws["J14"].value = info.get('fecha', "")
    ws["J15"].value = info.get('celular', "")

    # 2) Fila base
    base_row = 21

    # 3) Insertar filas extras si hay más de 1 producto
    if len(productos) > 1:
        for _ in range(len(productos) - 1):
            ws.insert_rows(base_row + 1)

    # Fila problemática especial: fila 22 (segunda fila de productos)
    try:
        # Eliminar cualquier merge existente en la fila 22
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row == 22 and rng.max_row == 22:
                ws.unmerge_cells(str(rng))
    except Exception as e:
        print(f"Warning al deshacer merges en fila 22: {e}")

    # Aplicar merge correcto para descripción (C-D) y marca/modelo (K-L)
    ws.merge_cells(start_row=22, start_column=3, end_row=22, end_column=4)   # C-D
    ws.merge_cells(start_row=22, start_column=11, end_row=22, end_column=12) # K-L

    # 4) Copiar estilos de la fila base a todas las filas de productos
    for i in range(base_row, base_row + len(productos)):
        for col in range(2, 13):  # B..L => 2..12
            src_cell = ws.cell(row=base_row, column=col)
            dst_cell = ws.cell(row=i, column=col)
            dst_cell.font = copy(src_cell.font)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)

        ws.row_dimensions[i].height = ws.row_dimensions[base_row].height


    # 4.5) Asegurarse de que cada fila tenga merges correctos en C-D y K-L
    for i in range(base_row-1, base_row + len(productos)):
        # Desfusionar C-D y K-L por si ya están fusionadas
        try:
            ws.unmerge_cells(start_row=i, start_column=3, end_row=i, end_column=4)
        except:
            pass
        try:
            ws.unmerge_cells(start_row=i, start_column=11, end_row=i, end_column=12)
        except:
            pass

        # Aplicar merge en C-D (col 3-4) y en K-L (col 11-12)
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=4)
        ws.merge_cells(start_row=i, start_column=11, end_row=i, end_column=12)

    tipo_cambio = info.get("tipo_cambio", "Soles").lower()
    valor_cambio = float(info.get("valor_cambio", 1.0))
    moneda_texto = tipo_cambio.upper()

    # 5) Rellenar los productos, incluyendo merges en C..D y K..L
    for idx, product in enumerate(productos):
        row_i = base_row + idx

        # 1. Obtener precios y aplicar conversiones
        precio_unitario = product.get("precio_unitario", 0.0)
        precio_total = product.get("precio_total", 0.0)

        # Sin IGV (18%) y convertir si no es en soles
        precio_unitario_sin_igv = precio_unitario / 1.18
        precio_total_sin_igv = precio_total / 1.18

        if tipo_cambio != "soles":
            precio_unitario_sin_igv /= valor_cambio
            precio_total_sin_igv /= valor_cambio

        # Redondear a 2 decimales
        precio_unitario_sin_igv = round(precio_unitario_sin_igv, 2)
        precio_total_sin_igv = round(precio_total_sin_igv, 2)

        # Ítem
        ws.cell(row=row_i, column=2).value = idx + 1
        ws.merge_cells(start_row=row_i, start_column=3, end_row=row_i, end_column=4)
        ws.cell(row=row_i, column=3).value = product.get("nombre_producto", "")
        ws.cell(row=row_i, column=5).value = product.get("unidad", "")
        ws.cell(row=row_i, column=6).value = product.get("cantidad", 0)
        ws.cell(row=row_i, column=8).value = precio_unitario_sin_igv
        ws.cell(row=row_i, column=9).value = precio_total_sin_igv
        ws.merge_cells(start_row=row_i, start_column=11, end_row=row_i, end_column=12)
        ws.cell(row=row_i, column=11).value = product.get("marca_modelo", "")

    # 6) Fila de suma => base_row + len(productos)
    sum_row = base_row + len(productos)

    # Copiar estilos a la fila de total
    for col in range(2, 13):
        src = ws.cell(row=base_row, column=col)
        dst = ws.cell(row=sum_row, column=col)
        dst.font = copy(src.font)
        dst.alignment = copy(src.alignment)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)

    # 1. Establecer altura de la fila total
    ws.row_dimensions[sum_row].height = ws.row_dimensions[base_row].height

    # 2. Desfusionar toda la fila por si acaso
    for start_col, end_col in [(3, 4), (11, 12), (2, 6), (3, 6)]:
        try:
            ws.unmerge_cells(start_row=sum_row, start_column=start_col, end_row=sum_row, end_column=end_col)
        except:
            pass

    # 3. Fondo blanco en toda la fila total
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for col in range(2, 13):  # B..L
        cell = ws.cell(row=sum_row, column=col)
        cell.fill = white_fill

    # 4. Merge de C..F para el texto
    ws.merge_cells(start_row=sum_row, start_column=3, end_row=sum_row, end_column=6)
    cell_label = ws.cell(row=sum_row, column=3)
    cell_label.value = "COSTO TOTAL SIN IGV EN SOLES S/."
    cell_label.font = Font(name="Helvetica Neue", size=11, color="000000", bold=True)
    cell_label.alignment = Alignment(horizontal="right", vertical="center")

    # 5. Celda I => el total con fórmula
    cell_total = ws.cell(row=sum_row, column=9)  # Columna I
    total_final = sum([product.get("precio_total", 0.0) / 1.18 / valor_cambio for product in productos])
    cell_total.value = round(total_final, 2)

    tipo_moneda = info.get("tipo_cambio", "Soles") 

    if tipo_moneda.lower() == "soles":
        label_text = "COSTO TOTAL SIN IGV EN SOLES S/."
    elif tipo_moneda.lower() == "dólares":
        label_text = "COSTO TOTAL SIN IGV EN DÓLARES $"
    elif tipo_moneda.lower() == "euros":
        label_text = "COSTO TOTAL SIN IGV EN EUROS €"
    else:
        label_text = "COSTO TOTAL SIN IGV"

    cell_label.value = label_text

    # Como la suma fue transformada antes, usamos sum directa:
    suma_total = sum(round(product.get("precio_total", 0.0) / 1.18 / valor_cambio, 2) for product in productos)
    cell_total.value = round(suma_total, 2)
    cell_total.font = Font(name="Helvetica Neue", size=11, color="000000", bold=True)
    cell_total.alignment = Alignment(horizontal="center", vertical="center")


    # 7) Insertar la sección final (footer) 2 filas debajo
    footer_row = sum_row + 2
    insert_footer_section(ws, footer_row, info)  # Tu función con merges y estilos

    # 8) Guardar y retornar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def insert_footer_section(ws, start_row, info=None):
    """
    Crea manualmente:
      - Fila verde 'Condiciones Comerciales' (con bordes y letra blanca).
      - 3 líneas debajo sin bordes (Costos, Entrega, Plazo).
      - Fila verde 'Información Comercial' (con bordes y letra blanca).
      - Tabla con RUC, Razón Social, etc. con bordes y letra negra.
      - Ajusta altura de todas las filas a 30.
    """
    if info is None:
        info = {}

    # STYLES
    green_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    white_font_11 = Font(name="Helvetica Neue", size=11, color="FFFFFFFF", bold=True)
    black_font_11 = Font(name="Helvetica Neue", size=11, color="FF000000")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # 1) Fila verde "Condiciones Comerciales" => B..K => con bordes, letra blanca
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=11)
    cond_cell = ws.cell(row=start_row, column=2)
    cond_cell.value = "Condiciones Comerciales:"
    cond_cell.fill = green_fill
    cond_cell.font = white_font_11
    cond_cell.alignment = center_align

    # Aplicar bordes en esa fila (B..K)
    for col in range(2, 12):
        cell = ws.cell(row=start_row, column=col)
        cell.border = thin_border

    # 2) Filas con texto sin bordes
    row_cc1 = start_row + 1
    row_cc2 = start_row + 2
    row_cc3 = start_row + 3

    entrega = info.get("lugar_entrega", "Moquegua")
    plazo = info.get("plazo_entrega", 1)

    ws.cell(row=row_cc1, column=3).value = "Costos no incluyen IGV"
    ws.cell(row=row_cc1, column=3).font = black_font_11
    ws.cell(row=row_cc1, column=3).alignment = left_align

    ws.cell(row=row_cc2, column=3).value = f"Entrega {entrega}"
    ws.cell(row=row_cc2, column=3).font = black_font_11
    ws.cell(row=row_cc2, column=3).alignment = left_align

    ws.cell(row=row_cc3, column=3).value = f"Plazo de entrega {plazo:02d} día"
    ws.cell(row=row_cc3, column=3).font = black_font_11
    ws.cell(row=row_cc3, column=3).alignment = left_align

    # 3) Fila verde “Información Comercial” => 2 filas debajo
    info_title_row = row_cc3 + 2
    ws.merge_cells(start_row=info_title_row, start_column=2, end_row=info_title_row, end_column=11)
    info_cell = ws.cell(row=info_title_row, column=2)
    info_cell.value = "Información Comercial"
    info_cell.fill = green_fill
    info_cell.font = white_font_11
    info_cell.alignment = center_align

    # Bordes en esa fila
    for col in range(2, 12):
        cell = ws.cell(row=info_title_row, column=col)
        cell.border = thin_border

    # 4) Tabla de Info Comercial (bordes + letra negra)
    row_rsocial = info_title_row + 1

    # Razón Social
    ws.merge_cells(start_row=row_rsocial, start_column=2, end_row=row_rsocial, end_column=3)
    ws.cell(row=row_rsocial, column=2).value = "Razón Social"
    ws.cell(row=row_rsocial, column=4).value = ": SEGURIMAX PERU SAC"
    ws.merge_cells(start_row=row_rsocial, start_column=5, end_row=row_rsocial, end_column=11)

    # RUC
    row_ruc = row_rsocial + 1
    ws.merge_cells(start_row=row_ruc, start_column=2, end_row=row_ruc, end_column=3)
    ws.cell(row=row_ruc, column=2).value = "RUC"
    ws.cell(row=row_ruc, column=4).value = ": 20604326916"
    ws.merge_cells(start_row=row_ruc, start_column=5, end_row=row_ruc, end_column=6)
    ws.cell(row=row_ruc, column=5).value = "Dirección"
    ws.merge_cells(start_row=row_ruc, start_column=7, end_row=row_ruc, end_column=11)
    ws.cell(row=row_ruc, column=7).value = ": Urb. Villa Hermosa I-10 Moquegua"

    # Cuentas
    row_bcp = row_ruc + 1
    ws.merge_cells(start_row=row_bcp, start_column=2, end_row=row_bcp, end_column=3)
    ws.cell(row=row_bcp, column=2).value = "Cuenta Soles BCP"
    ws.cell(row=row_bcp, column=4).value = ": 430-2584079-0-40"
    ws.merge_cells(start_row=row_bcp, start_column=5, end_row=row_bcp, end_column=6)
    ws.cell(row=row_bcp, column=5).value = "Cuenta Dólares BCP"
    ws.merge_cells(start_row=row_bcp, start_column=7, end_row=row_bcp, end_column=11)
    ws.cell(row=row_bcp, column=7).value = ": 430-2609389-1-07"

    # Contacto
    row_contacto = row_bcp + 1
    ws.merge_cells(start_row=row_contacto, start_column=2, end_row=row_contacto, end_column=3)
    ws.cell(row=row_contacto, column=2).value = "Contacto"
    ws.cell(row=row_contacto, column=4).value = ": Maria Luisa García Núñez"
    ws.merge_cells(start_row=row_contacto, start_column=5, end_row=row_contacto, end_column=6)
    ws.cell(row=row_contacto, column=5).value = "Celular"
    ws.merge_cells(start_row=row_contacto, start_column=7, end_row=row_contacto, end_column=11)
    ws.cell(row=row_contacto, column=7).value = ": 996665221"

    # Email
    row_email = row_contacto + 1
    ws.merge_cells(start_row=row_email, start_column=2, end_row=row_email, end_column=3)
    ws.cell(row=row_email, column=2).value = "Email"
    ws.cell(row=row_email, column=4).value = ": segurimaxperu1@gmail.com"
    ws.merge_cells(start_row=row_email, start_column=5, end_row=row_email, end_column=11)

    # 5) Poner bordes + font en la tabla (fila info_title_row + 1 .. row_email)
    #    y en cada celda => black font, left align (excepto merges con text?)
    #    Ajustamos la altura = 30 en TODAS las filas del footer
    last_row_used = row_email
    for row in range(info_title_row+1, last_row_used+1):
        for col in range(2, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = black_font_11
            cell.alignment = left_align

    # Además, ajustamos la altura de TODAS las filas, incluso la verde
    for row in range(start_row, last_row_used+1):
        ws.row_dimensions[row].height = 30

    return last_row_used

